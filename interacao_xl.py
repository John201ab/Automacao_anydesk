# le linha por linha da planilha
def dados():
    import pyautogui as pa
    from openpyxl import load_workbook
    import time

    planilha_resultado = load_workbook('resultado.xlsx')
    pg_principal = planilha_resultado['Sheet1']

    while True:
        for linhas in pg_principal.iter_rows(values_only=True):
            for loop in range(2):
                celula = linhas[loop]
                if loop == 1:
                    time.sleep(2)
                    pa.click(x=-1130, y=481)
                    pa.write(str(celula))
                if loop == 0:
                    pa.click(x=-1154, y=451)
                    time.sleep(1)
                    pa.hotkey('ctrl','a')
                    time.sleep(1)
                    pa.press('backspace')
                    time.sleep(2)
                    pa.click(x=-1154, y=451)
                    pa.write(str(celula))
                    time.sleep(2)
                    pa.click(x=-1130, y=481)
                    pa.hotkey('ctrl','a')
                    pa.press('backspace')
            pa.click(x=-1151, y=538)
        break